"""
World Bank Global Fuel Prices Database Extraction Script
========================================================
Extracts monthly fuel prices for 8 Latin American countries
from the World Bank Excel dataset and outputs a CSV file
suitable for bulk upsert into fuel_prices_regional.

Products: Regular Gasoline, Premium Gasoline, Diesel, Kerosene
Countries: Argentina, Brazil, Chile, Colombia, Ecuador,
           Mexico, Peru, Venezuela (RB)
Note: Paraguay has no product data in the WB dataset.

Usage: python extract-wb-data.py [excel_file_path] [output_csv_path]
"""

import openpyxl
import csv
import sys
import os
from datetime import datetime, date
from collections import defaultdict

# ── Configuration ──────────────────────────────────────────────────────────
EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\nicol\AppData\Local\Temp\opencode\Global_Fuel_Prices_Database.xlsx'
OUTPUT_CSV = sys.argv[2] if len(sys.argv) > 2 else r'C:\Users\nicol\AppData\Local\Temp\opencode\wb_fuel_prices.csv'

# Country name mapping: WB → Dashboard
COUNTRY_MAP = {
    'Argentina': 'Argentina',
    'Brazil': 'Brasil',
    'Chile': 'Chile',
    'Colombia': 'Colombia',
    'Ecuador': 'Ecuador',
    'Mexico': 'México',
    'Peru': 'Perú',
    'Venezuela, RB': 'Venezuela',
    'Venezuela': 'Venezuela',
}

# Product name mapping: WB product sheet → dashboard product name
PRODUCT_MAP = {
    'regular': 'Gasolina Regular',
    'premium': 'Gasolina Premium',
    'diesel': 'Diésel',
    'kerosene': 'Kerosene',
}

# Product sheets in the WB Excel
PRODUCT_CONFIG = {
    'regular': {
        'local_sheet': 'Regular Gasoline (below RON 95)',
        'usd_sheet': 'Reg Gasoline (below RON 95) USD',
    },
    'premium': {
        'local_sheet': 'Premium Gasoline RON95 or above',
        'usd_sheet': 'Premium Gasoline RON95or ab USD',
    },
    'diesel': {
        'local_sheet': None,  # Diesel sheet has different structure
        'usd_sheet': 'Diesel USD',
    },
    'kerosene': {
        'local_sheet': 'Kerosene',
        'usd_sheet': 'Kerosene USD',
    },
}

# Fixed currency codes per country
CURRENCY_MAP = {
    'Argentina': 'ARS',
    'Brasil': 'BRL',
    'Chile': 'CLP',
    'Colombia': 'COP',
    'Ecuador': 'USD',
    'México': 'MXN',
    'Perú': 'PEN',
    'Venezuela': 'VES',
}

# Price validation bounds (USD/liter)
BOUNDS = {
    'Gasolina Regular': (0.01, 3.00),
    'Gasolina Premium': (0.01, 3.00),
    'Diésel': (0.001, 3.00),
    'Kerosene': (0.01, 3.00),
}

TARGET_COUNTRIES_WB = ['Argentina', 'Brazil', 'Chile', 'Colombia', 'Ecuador', 'Mexico', 'Peru', 'Venezuela']

# ── Helpers ────────────────────────────────────────────────────────────────

def parse_excel_date(val):
    """Parse datetime from Excel cell value (could be datetime or string)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        # Try to parse as date
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def validate(record):
    """Validate a record against business rules. Returns (is_valid, reason)."""
    pais = record.get('pais', '')
    producto = record.get('producto', '')
    
    # Check nulls
    if not record.get('fecha'):
        return False, 'null fecha'
    if record.get('precio_usd') is None:
        return False, 'null precio_usd'
    
    # Parse precio_usd
    try:
        precio = float(record['precio_usd'])
    except (ValueError, TypeError):
        return False, 'non-numeric precio_usd'
    
    if precio <= 0:
        return False, 'zero or negative price'
    
    # Price bounds
    bounds = BOUNDS.get(producto, (0.001, 3.00))
    if precio < bounds[0] or precio > bounds[1]:
        # Venezuela special case: diesel can be as low as $0.001
        if pais == 'Venezuela' and bounds[0] <= precio <= bounds[1]:
            pass  # Accept Venezuelan subsidized prices
        else:
            return False, f'price {precio} out of range [{bounds[0]}, {bounds[1]}]'
    
    # Date validation
    fecha = record['fecha']
    if isinstance(fecha, str):
        try:
            fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
        except ValueError:
            return False, f'invalid date format: {fecha}'
    
    if fecha > date.today():
        return False, f'future date: {fecha}'
    
    return True, 'ok'


def spin_wait(ms):
    """Busy-wait for ms milliseconds (for edge function compatibility)."""
    import time
    start = time.time()
    while (time.time() - start) * 1000 < ms:
        pass


# ── Main Extraction ────────────────────────────────────────────────────────

def load_exchange_rates(ws_er):
    """Load exchange rates from ER sheet.
    Returns: dict[country_name][date_str] = rate (LCU per USD)
    """
    rates = defaultdict(dict)
    
    # Parse header row for dates (cols 3+)
    dates = []
    for c in range(3, ws_er.max_column + 1):
        header = ws_er.cell(row=1, column=c).value
        d = parse_excel_date(header)
        if d:
            dates.append((c, d.isoformat()))
    
    # Parse each country row
    for r in range(2, ws_er.max_row + 1):
        country_name = str(ws_er.cell(row=r, column=1).value or '').strip()
        # Map WB name to mapped name
        mapped_country = COUNTRY_MAP.get(country_name)
        if not mapped_country:
            # Also try partial match
            for wb_name, mapped in COUNTRY_MAP.items():
                if wb_name.lower() in country_name.lower():
                    mapped_country = mapped
                    break
        
        if mapped_country and mapped_country not in [COUNTRY_MAP.get(c) for c in TARGET_COUNTRIES_WB]:
            continue
        if not mapped_country:
            continue
        
        for col, date_str in dates:
            val = ws_er.cell(row=r, column=col).value
            if val is not None:
                try:
                    rate = float(val)
                    if rate > 0:
                        rates[mapped_country][date_str] = rate
                except (ValueError, TypeError):
                    pass
    
    return rates


def extract_product_data(wb, product_key, config):
    """Extract data for a single product from USD sheet.
    Returns list of record dicts.
    """
    usd_sheet_name = config['usd_sheet']
    if usd_sheet_name not in wb.sheetnames:
        print(f'  WARNING: Sheet "{usd_sheet_name}" not found, skipping {product_key}')
        return []
    
    ws = wb[usd_sheet_name]
    records = []
    stats = {'found': 0, 'rows': 0, 'valid': 0, 'rejected': 0}
    
    # Parse header row for dates (cols 7+, starting after Default MAP col)
    # USD sheet structure: Col1=LongName, Col2=Country, Col3=Code, Col4=OrigUnits, 
    #                     Col5=ConvUnits, Col6=DefaultMAP, Col7+=dates
    date_cols = []
    # Scan ALL columns for date headers (starting from col 3+ to avoid title cols)
    for c in range(3, ws.max_column + 1):
        header_val = ws.cell(row=1, column=c).value
        d = parse_excel_date(header_val)
        if d:
            date_cols.append((c, d.isoformat()))
    
    # Fallback: if no dates found, try starting at col 7
    if not date_cols:
        for c in range(7, ws.max_column + 1):
            header_val = ws.cell(row=1, column=c).value
            d = parse_excel_date(header_val)
            if d:
                date_cols.append((c, d.isoformat()))
    
    print(f'  Product: {PRODUCT_MAP[product_key]} ({product_key})')
    print(f'    Sheet: {usd_sheet_name}, {ws.max_row} rows, date cols: {len(date_cols)}')
    
    # Find rows for each target country, prefer MAP=1
    # store all candidate rows per country: (row_num, country_name, is_map1)
    country_candidates = defaultdict(list)
    
    for r in range(2, ws.max_row + 1):
        country_name = str(ws.cell(row=r, column=2).value or '').strip()
        
        # Check if this is a target country
        is_target = False
        mapped_name = None
        for wb_name in TARGET_COUNTRIES_WB:
            if country_name == wb_name or wb_name in country_name:
                is_target = True
                mapped_name = COUNTRY_MAP.get(wb_name, country_name)
                break
        
        if not is_target:
            # Try Venezuela, RB
            if 'Venezuela' in country_name:
                is_target = True
                mapped_name = 'Venezuela'
        
        if not is_target:
            continue
        
        map_flag = ws.cell(row=r, column=6).value  # Default MAP column
        is_map1 = (map_flag == 1 or str(map_flag).strip() == '1')
        
        country_candidates[mapped_name].append((r, country_name, is_map1))
    
    # Select best row: MAP=1 preferred; if MAP=1 has no data, try MAP=0
    country_best_rows = {}
    for mapped_name, candidates in country_candidates.items():
        # Try MAP=1 first
        map1_rows = [(rn, cn, m) for rn, cn, m in candidates if m]
        map0_rows = [(rn, cn, m) for rn, cn, m in candidates if not m]
        
        # Check if MAP=1 row has data in the date range
        selected = None
        for rn, cn, m in map1_rows:
            # Sample first 5 date columns to check for data
            has_data = False
            for col, _ in date_cols[:5]:
                val = ws.cell(row=rn, column=col).value
                if val is not None:
                    try:
                        if float(val) > 0:
                            has_data = True
                            break
                    except (ValueError, TypeError):
                        pass
            if has_data:
                selected = (rn, cn, True)
                break
        
        # If no MAP=1 row has data, fall back to MAP=0
        if not selected and map0_rows:
            selected = (map0_rows[0][0], map0_rows[0][1], False)
        elif not selected:
            # Last resort: any candidate
            selected = (candidates[0][0], candidates[0][1], candidates[0][2])
        
        if selected:
            country_best_rows[mapped_name] = selected
    
    stats['found'] = len(country_best_rows)
    print(f'    Countries with data: {list(country_best_rows.keys())}')
    print(f'    MAP=1 rows: {sum(1 for _,_,m in country_best_rows.values() if m)}')
    
    # Extract data from best rows
    for mapped_name, (row_num, wb_country, is_map1) in country_best_rows.items():
        country_records = 0
        for col, date_str in date_cols:
            usd_val = ws.cell(row=row_num, column=col).value
            if usd_val is None:
                continue
            
            try:
                precio_usd = round(float(usd_val), 6)
            except (ValueError, TypeError):
                continue
            
            if precio_usd <= 0:
                continue
            
            record = {
                'fecha': date_str,
                'pais': mapped_name,
                'producto': PRODUCT_MAP[product_key],
                'precio_usd': precio_usd,
                'precio_local': 0,  # Will be filled later
                'moneda': CURRENCY_MAP.get(mapped_name, 'USD'),
                'fuente': 'World Bank',
                'data_source': 'WB Global Fuel Prices DB',
                'frecuencia': 'mensual',
                'quality_score': 1.0,
            }
            
            is_valid, reason = validate(record)
            if is_valid:
                records.append(record)
                stats['valid'] += 1
                country_records += 1
            else:
                stats['rejected'] += 1
        
        print(f'    {mapped_name}: {country_records} valid records')
    
    stats['rows'] = len(records)
    print(f'    Total: {stats["valid"]} valid, {stats["rejected"]} rejected')
    return records


def fill_local_prices(records, rates):
    """Calculate precio_local = precio_usd × exchange_rate for each record.
    Missing rates are backfilled with the closest prior rate."""
    filled = 0
    missing_rate = 0
    for rec in records:
        pais = rec['pais']
        fecha = rec['fecha']
        
        rate = None
        
        # Try exact date match first
        if pais in rates and fecha in rates[pais]:
            rate = rates[pais][fecha]
        
        # Try year-month match (different day)
        if not rate and pais in rates:
            pais_rates = rates[pais]
            for rate_date, r in pais_rates.items():
                if rate_date[:7] == fecha[:7]:
                    rate = r
                    break
        
        # Backfill: find closest prior rate
        if not rate and pais in rates:
            pais_rates = rates[pais]
            sorted_dates = sorted(pais_rates.keys())
            best_date = None
            for rd in sorted_dates:
                if rd < fecha:
                    best_date = rd
                else:
                    break
            if best_date:
                rate = pais_rates[best_date]
        
        # Forward fill: if no prior rate, use earliest available
        if not rate and pais in rates:
            pais_rates = rates[pais]
            if pais_rates:
                sorted_dates = sorted(pais_rates.keys())
                rate = pais_rates[sorted_dates[0]]
        
        if rate and rate > 0:
            rec['precio_local'] = round(rec['precio_usd'] * rate, 4)
            filled += 1
        else:
            # Last resort: use precio_usd as precio_local (for dollarized economies)
            if pais == 'Ecuador' or pais == 'Venezuela':
                rec['precio_local'] = rec['precio_usd']
                rec['moneda'] = 'USD' if pais == 'Ecuador' else 'VES'
                filled += 1
            else:
                missing_rate += 1
    
    print(f'  Local prices filled: {filled}, missing rates: {missing_rate}')
    return records


def main():
    print('=' * 70)
    print('World Bank Global Fuel Prices Database — Extraction')
    print('=' * 70)
    
    # Load workbook
    print(f'\nLoading Excel: {EXCEL_PATH}')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f'Sheets: {wb.sheetnames[:5]}... ({len(wb.sheetnames)} total)')
    
    # Load exchange rates
    print('\n--- Exchange Rates ---')
    er_sheet_name = None
    for sn in wb.sheetnames:
        if 'Monthly ER' in sn and 'FORMULA' not in sn:
            er_sheet_name = sn
            break
    
    rates = {}
    if er_sheet_name:
        print(f'Loading ER from: {er_sheet_name}')
        rates = load_exchange_rates(wb[er_sheet_name])
        print(f'Countries with rates: {list(rates.keys())}')
        for country, country_rates in rates.items():
            print(f'  {country}: {len(country_rates)} monthly rates')
    
    # Extract product data
    all_records = []
    total_stats = {'valid': 0, 'rejected': 0, 'products': 0}
    
    for product_key, config in PRODUCT_CONFIG.items():
        print(f'\n--- {product_key.upper()} ---')
        records = extract_product_data(wb, product_key, config)
        if records:
            all_records.extend(records)
            total_stats['products'] += 1
    
    # Fill local prices
    print('\n--- Filling Local Prices ---')
    all_records = fill_local_prices(all_records, rates)
    
    # Summary
    print('\n' + '=' * 70)
    print('EXTRACTION SUMMARY')
    print('=' * 70)
    countries = set(r['pais'] for r in all_records)
    products = set(r['producto'] for r in all_records)
    print(f'Total records: {len(all_records)}')
    print(f'Countries: {sorted(countries)} ({len(countries)} total)')
    print(f'Products: {sorted(products)}')
    
    for c in sorted(countries):
        c_records = [r for r in all_records if r['pais'] == c]
        c_products = set(r['producto'] for r in c_records)
        print(f'  {c}: {len(c_records)} records, products: {sorted(c_products)}')
    
    # Write CSV
    if all_records:
        fieldnames = ['fecha', 'pais', 'producto', 'precio_usd', 'precio_local', 
                      'moneda', 'fuente', 'data_source', 'frecuencia', 'quality_score']
        
        with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_records)
        
        print(f'\nCSV written: {OUTPUT_CSV}')
        print(f'File size: {os.path.getsize(OUTPUT_CSV):,} bytes')
    else:
        print('\nERROR: No records extracted!')
    
    # Show sample
    if all_records:
        print('\n--- Sample Records (first 5) ---')
        for r in all_records[:5]:
            print(f'  {r["fecha"]} | {r["pais"]:12s} | {r["producto"]:20s} | USD {r["precio_usd"]:.4f} | {r["moneda"]} {r["precio_local"]:.4f}')
    
    return 0 if all_records else 1


if __name__ == '__main__':
    sys.exit(main())
