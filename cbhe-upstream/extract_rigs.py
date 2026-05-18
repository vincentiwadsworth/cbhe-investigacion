"""
Extract rig counts from Baker Hughes xlsx files and update InsForge upstream_annual table.
Usage: python extract_rigs.py
"""
import openpyxl
import json
import os
import re
from collections import defaultdict

BASE_DIR = r"C:\Users\nicol\Documents\GitHub\cbhe-investigacion\cbhe-upstream"

# Countries we care about (Baker Hughes name -> our name)
COUNTRY_MAP = {
    "ARGENTINA": "Argentina",
    "BOLIVIA": "Bolivia",
    "BRAZIL": "Brasil",
    "CHILE": "Chile",
    "COLOMBIA": "Colombia",
    "ECUADOR": "Ecuador",
    "MEXICO": "México",
    "PERU": "Perú",
    "VENEZUELA": "Venezuela",
}

files = [
    os.path.join(BASE_DIR, "July-2025 WorldWide Rig Count Report.xlsx"),
    os.path.join(BASE_DIR, "April-2026 WorldWide Rig Count Report.xlsx"),
]

# Collect all monthly data points: {country_name: {year: [monthly_counts]}}
rig_data = defaultdict(lambda: defaultdict(list))

for filepath in files:
    print(f"\nProcessing: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column})")
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 7:
                continue
            
            region = row[0]
            country_raw = row[1]
            
            if not country_raw or not isinstance(country_raw, str):
                continue
            
            country_key = country_raw.strip().upper()
            
            if country_key not in COUNTRY_MAP:
                continue
            
            country_name = COUNTRY_MAP[country_key]
            rig_type = row[2]  # Oil, Gas, Miscellaneous
            location = row[3]  # Land, Offshore, etc
            year = row[4]
            month = row[5]
            count = row[6]
            
            # Also check columns 7+ for historical data
            # But first handle the current data point
            if year and month and count is not None:
                try:
                    y = int(year)
                    c = float(count)
                    rig_data[country_name][y].append(c)
                except (ValueError, TypeError):
                    pass
            
            # Columns 7+ may contain additional historical months
            for i in range(7, len(row)):
                val = row[i]
                if val is not None:
                    try:
                        c = float(val)
                        # Need to figure out the year/month for this column
                        # The column headers in row 1 should tell us
                        # For now, skip these - we'll use the primary data
                    except (ValueError, TypeError):
                        pass

# Now aggregate: average rigs per country per year
print("\n\n=== AGGREGATED RIG COUNTS (annual average) ===")
results = {}
for country in sorted(rig_data.keys()):
    results[country] = {}
    for year in sorted(rig_data[country].keys()):
        counts = rig_data[country][year]
        if counts:
            # Sum all types for the year, then average by number of months
            # Each entry is one type+location for one month
            # We need to group by month first, sum across types, then average
            pass
    
# Actually, let me recalculate. Each row is one type+location for one specific month.
# So for Argentina in 2025 month 6: Gas Land=12, Oil Land=30 = 42 total for that month
# For Argentina in 2025 month 7: Gas Land=11, Oil Land=27 = 38 total for that month
# Annual average = mean of monthly totals

# Better approach: group by country -> year -> month -> sum across types
monthly_totals = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
all_entries = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

for filepath in files:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 7:
                continue
            country_raw = row[1]
            if not country_raw or not isinstance(country_raw, str):
                continue
            country_key = country_raw.strip().upper()
            if country_key not in COUNTRY_MAP:
                continue
            country_name = COUNTRY_MAP[country_key]
            year = row[4]
            month = row[5]
            count = row[6]
            if year and month and count is not None:
                try:
                    y = int(year)
                    m = int(month)
                    c = float(count)
                    all_entries[country_name][y][m] = all_entries[country_name][y].get(m, 0) + c
                except (ValueError, TypeError):
                    pass

# Calculate annual averages
print("\n=== RIG COUNTS BY COUNTRY (annual avg of monthly totals) ===")
output_rows = []
for country in sorted(all_entries.keys()):
    for year in sorted(all_entries[country].keys()):
        months = all_entries[country][year]
        if months:
            avg = sum(months.values()) / len(months)
            avg_rounded = round(avg, 1)
            print(f"  {country} {year}: avg={avg_rounded} (from {len(months)} months)")
            output_rows.append({
                "pais": country,
                "anio": year,
                "rigs_activos_prom": avg_rounded,
                "months_data": len(months),
                "monthly_values": {str(m): round(v, 1) for m, v in sorted(months.items())}
            })

# Save to JSON for inspection
output_path = os.path.join(BASE_DIR, "rig_counts_extracted.json")
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(output_rows, f, indent=2, ensure_ascii=False)
print(f"\nSaved to: {output_path}")
print(f"Total rows: {len(output_rows)}")
